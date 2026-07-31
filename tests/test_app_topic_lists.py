"""Editing a topic list from the app.

The editor writes the same spreadsheet the run reads, so what it accepts and what the run
accepts have to be the same thing — these tests pin that, and the "not a set until you name it"
rule that keeps half-finished lists out of a tagging run.
"""
import pytest

from transcript_toolkit.app import topic_lists
from transcript_toolkit.errors import ToolkitError
from transcript_toolkit.project import init_project


@pytest.fixture
def project(tmp_path):
    return init_project(str(tmp_path / "ws"))


ROWS = [{"id": "", "name": "Education", "description": "Schooling and universities."},
        {"id": "", "name": "Career and Work", "description": "Jobs and organizations."}]


def test_the_shipped_example_is_what_you_start_from(project):
    """The example file is the starting point, and its instruction row is guidance, not a topic
    somebody has to notice and delete."""
    rows, guidance = topic_lists.load_rows(topic_lists.draft_path(project))
    assert guidance and "RENAME this file" in guidance
    assert all(r["id"] != topic_lists.HOW_TO_ROW for r in rows)
    assert [r["name"] for r in rows][:1] == ["Education"]


def test_a_draft_is_kept_but_is_not_a_set_anybody_can_run(project):
    """Autosave must never turn half-written topics into something a run will tag against."""
    from transcript_toolkit.steps.topics.taxonomy import discover_topic_files

    topic_lists.save_draft(project, [{"id": "", "name": "Half a t", "description": ""}])
    assert topic_lists.draft_path(project).exists()
    assert discover_topic_files(project) == {}


def test_naming_a_list_writes_it_and_registers_it(project):
    from transcript_toolkit.core.config import load_root_config

    set_name, path = topic_lists.save_as(project, "Collection", ROWS)
    assert set_name == "collection" and path.name == "collection.csv"
    sets = (load_root_config(project).get("topics") or {}).get("sets") or {}
    assert sets["collection"]["file"] == "topics/collection.csv"

    # and the step can now load exactly what was typed
    from transcript_toolkit.core.config import load_step_config
    from transcript_toolkit.steps.topics.taxonomy import load_topic_set
    tset = load_topic_set(project, load_step_config(project, "topics"), "collection")
    assert tset.ids == ["education", "career_and_work"]
    assert "## Education\n\nSchooling and universities." in tset.taxonomy_text


def test_the_editor_refuses_what_the_run_would_refuse_in_the_same_words(project):
    """A list the editor accepted and the run then rejected would be the worst of both."""
    with pytest.raises(ToolkitError, match="empty description"):
        topic_lists.save_as(project, "bad", [{"id": "", "name": "Nameless", "description": ""}])
    with pytest.raises(ToolkitError, match="duplicate topic id"):
        topic_lists.save_as(project, "dupes", [
            {"id": "x", "name": "One", "description": "a"},
            {"id": "x", "name": "Two", "description": "b"}])
    assert not topic_lists.set_path(project, "bad").exists()


def test_a_name_that_is_not_a_filename_is_turned_into_one(project):
    set_name, path = topic_lists.save_as(project, "Main Themes!", ROWS)
    assert set_name == "main_themes" and path.name == "main_themes.csv"
    with pytest.raises(ToolkitError, match="cannot be a topic list name"):
        topic_lists.save_as(project, "!!!", ROWS)


def test_an_existing_name_is_not_silently_overwritten(project):
    topic_lists.save_as(project, "collection", ROWS)
    with pytest.raises(ToolkitError, match="already a topic list"):
        topic_lists.save_as(project, "collection", ROWS)


def test_editing_a_named_list_writes_back_to_the_same_file(project):
    set_name, path = topic_lists.save_as(project, "collection", ROWS)
    topic_lists.save_existing(project, set_name, path,
                              ROWS + [{"id": "", "name": "Migration", "description": "Moving."}])
    rows, _ = topic_lists.load_rows(path)
    assert [r["name"] for r in rows][-1] == "Migration"


def test_an_excel_list_is_edited_as_an_excel_file(project):
    """A list somebody uploaded as a spreadsheet is editable here like any other, and stays the
    format they brought — rewriting it as .csv would leave two files claiming one set name."""
    from openpyxl import Workbook, load_workbook

    path = project.topics_dir / "collection.xlsx"
    book = Workbook()
    book.active.append(["name", "description"])
    book.active.append(["Old", "the row that was there"])
    book.create_sheet("notes").append(["something else of the user's"])
    book.save(path)

    topic_lists.save_existing(project, "collection", path, ROWS)

    assert path.suffix == ".xlsx"
    rows, _ = topic_lists.load_rows(path)
    assert [r["name"] for r in rows] == [r["name"] for r in ROWS]
    # and the rest of their workbook is still theirs
    assert "notes" in load_workbook(path).sheetnames


def test_the_set_the_editor_opens_is_the_file_that_was_uploaded(project):
    """An uploaded list has never been through the editor, so nothing has written a .csv for it —
    the page has to find the file that is actually there."""
    from transcript_toolkit.app.pages.step import set_file

    (project.topics_dir / "brought_along.xlsx").write_bytes(b"")
    assert set_file(project, "brought_along").name == "brought_along.xlsx"


def test_blank_rows_an_editor_collects_are_not_written(project):
    _, path = topic_lists.save_as(project, "collection",
                                  [*ROWS, {"id": "", "name": "", "description": ""}])
    rows, _ = topic_lists.load_rows(path)
    assert len(rows) == 2
