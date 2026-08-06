"""Hand-edited clip labels: the overrides file, the review-page edit, and the export loop.

A curator may fix one word of one label — in the review page or in the exported sheet — and
the pipeline must show that fix everywhere without ever writing it into the model's own
deliverable, and must drop it out loud when the clip it pinned no longer exists.
"""
import json

import pandas as pd
import pytest

from transcript_toolkit.core import overrides
from transcript_toolkit.errors import ToolkitError
from transcript_toolkit.project import init_project


@pytest.fixture
def project(tmp_path):
    return init_project(str(tmp_path / "ws"))


def clips_frame() -> pd.DataFrame:
    return pd.DataFrame([
        {"clip_id": "fake_beta_0001", "interview_id": "fake_beta",
         "start_ts": "00:00:05", "end_ts": "00:01:00", "start_paragraph_idx": 0,
         "duration_seconds": 55.0, "label": "Growing up by the sea"},
        {"clip_id": "fake_beta_0002", "interview_id": "fake_beta",
         "start_ts": "00:01:00", "end_ts": "00:02:00", "start_paragraph_idx": 2,
         "duration_seconds": 60.0, "label": "University years"},
    ])


# --- the store ------------------------------------------------------------------------------

def test_upsert_overlay_roundtrip(project):
    clips = clips_frame()
    overrides.upsert(project, "fake_beta_0001", "Growing up by the mountains", clips,
                     replaces="Growing up by the sea")
    shown, complaints = overrides.overlay(project, clips)
    assert shown == {"fake_beta_0001": "Growing up by the mountains"} and not complaints
    assert project.label_overrides_path.exists()      # a plain csv the CLI user can open

    assert overrides.remove(project, "fake_beta_0001")
    shown, _ = overrides.overlay(project, clips)
    assert shown == {}


def test_an_override_is_pinned_to_the_clips_span(project):
    clips = clips_frame()
    overrides.upsert(project, "fake_beta_0001", "My version", clips)

    moved = clips.copy()
    moved.loc[moved["clip_id"] == "fake_beta_0001", "end_ts"] = "00:01:30"
    shown, complaints = overrides.overlay(project, moved)
    assert shown == {} and len(complaints) == 1
    assert "span has changed" in complaints[0]

    gone = clips[clips["clip_id"] != "fake_beta_0001"]
    shown, complaints = overrides.overlay(project, gone)
    assert shown == {} and "no such clip" in complaints[0]


def test_a_reimported_interview_takes_its_overrides_with_it(project):
    clips = clips_frame()
    overrides.upsert(project, "fake_beta_0001", "My version", clips)
    assert overrides.purge_interviews(project, ["fake_beta"])
    assert overrides.load(project).empty
    assert not overrides.purge_interviews(project, ["fake_beta"])     # nothing left to drop


def test_an_empty_label_is_refused(project):
    with pytest.raises(ToolkitError, match="empty"):
        overrides.upsert(project, "fake_beta_0001", "   ", clips_frame())


# --- the export loop -------------------------------------------------------------------------

def fabricate_labels(project) -> None:
    out = project.outputs_dir
    (out / "clips").mkdir(parents=True, exist_ok=True)
    (out / "labels").mkdir(parents=True, exist_ok=True)
    clips = clips_frame()
    clips.drop(columns=["label"]).to_parquet(out / "clips" / "clips.parquet", index=False)
    clips.to_parquet(out / "labels" / "labels.parquet", index=False)


def test_export_shows_the_override_and_records_what_it_wrote(project):
    from transcript_toolkit.steps.export import run_export

    fabricate_labels(project)
    overrides.upsert(project, "fake_beta_0001", "Growing up by the mountains",
                     pd.read_parquet(project.outputs_dir / "labels" / "labels.parquet"),
                     replaces="Growing up by the sea")
    run_export(project)

    from openpyxl import load_workbook
    wb = load_workbook(project.outputs_dir / "export.xlsx", read_only=True)
    rows = list(wb["Clips"].iter_rows(values_only=True))
    header = list(rows[0])
    by_id = {r[header.index("Clip Id")]: r[header.index("Label")] for r in rows[1:]}
    assert by_id["fake_beta_0001"] == "Growing up by the mountains"
    assert by_id["fake_beta_0002"] == "University years"
    wb.close()

    manifest = json.loads(project.export_manifest_path.read_text())
    assert manifest["labels"]["fake_beta_0001"] == "Growing up by the mountains"


def test_a_label_edited_in_the_sheet_survives_the_next_export(project):
    """The whole point: edit one cell in Excel, run another step, re-export — the edit stays."""
    from openpyxl import load_workbook

    from transcript_toolkit.steps.export import run_export

    fabricate_labels(project)
    run_export(project)

    out = project.outputs_dir / "export.xlsx"
    wb = load_workbook(out)
    ws = wb["Clips"]
    header = [c.value for c in ws[1]]
    label_col = header.index("Label") + 1
    assert ws.cell(row=2, column=label_col).value == "Growing up by the sea"
    ws.cell(row=2, column=label_col).value = "Growing up by the shore"     # the curator's fix
    wb.save(out)
    wb.close()

    run_export(project)

    saved = overrides.load(project)
    assert list(saved["clip_id"]) == ["fake_beta_0001"]
    assert saved.iloc[0]["label"] == "Growing up by the shore"
    assert saved.iloc[0]["replaces"] == "Growing up by the sea"

    wb = load_workbook(out, read_only=True)
    rows = list(wb["Clips"].iter_rows(values_only=True))
    by_id = {r[0]: r[list(rows[0]).index("Label")] for r in rows[1:]}
    assert by_id["fake_beta_0001"] == "Growing up by the shore"
    wb.close()

    # edited back to the model's own words -> the override is removed again
    wb = load_workbook(out)
    ws = wb["Clips"]
    ws.cell(row=2, column=label_col).value = "Growing up by the sea"
    wb.save(out)
    wb.close()
    run_export(project)
    assert overrides.load(project).empty


# --- the review page -------------------------------------------------------------------------

def paras_frame() -> pd.DataFrame:
    return pd.DataFrame([
        {"interview_id": "fake_beta", "paragraph_idx": 0, "clip_id": "fake_beta_0001",
         "word_count": 5, "speaker_role": "Narrator", "speech": "By the sea.",
         "turn_time_start": "00:00:05", "sub_time_start": ""},
        {"interview_id": "fake_beta", "paragraph_idx": 2, "clip_id": "fake_beta_0002",
         "word_count": 4, "speaker_role": "Narrator", "speech": "Then university.",
         "turn_time_start": "00:01:00", "sub_time_start": ""},
    ])


def test_the_review_page_shows_the_edit_and_carries_the_edit_control(project):
    from transcript_toolkit.steps.label.annotate import write_annotated

    clips = clips_frame()
    overrides.upsert(project, "fake_beta_0001", "Growing up by the mountains", clips)
    label_by_id = {"fake_beta_0001": "Growing up by the sea",
                   "fake_beta_0002": "University years"}
    diag_dir = write_annotated(project, ["fake_beta"], paras_frame(),
                               clips.drop(columns=["label"]), label_by_id)

    page = (diag_dir / "fake_beta.html").read_text()
    assert 'data-clip="fake_beta_0001"' in page
    assert "Growing up by the mountains" in page and "edited by hand" in page
    assert "Growing up by the sea" not in page                 # the model's words are replaced
    assert "/api/labels/edit" in page                          # the edit control is in the page
    assert 'location.protocol === "file:"' in page             # and knows when to stand down


def test_the_edit_endpoint_patches_the_page_on_disk(project):
    from transcript_toolkit.app.server import _patch_label_page
    from transcript_toolkit.steps.label.annotate import write_annotated

    clips = clips_frame()
    label_by_id = {"fake_beta_0001": "Growing up by the sea",
                   "fake_beta_0002": "University years"}
    diag_dir = write_annotated(project, ["fake_beta"], paras_frame(),
                               clips.drop(columns=["label"]), label_by_id)

    _patch_label_page(project, "fake_beta_0001", "Growing up by the mountains", True)
    page = (diag_dir / "fake_beta.html").read_text()
    assert "Growing up by the mountains" in page and "edited by hand" in page
    assert page.count('data-clip="fake_beta_0001"') == 1       # replaced, not duplicated
    assert "University years" in page                          # the other clip untouched
