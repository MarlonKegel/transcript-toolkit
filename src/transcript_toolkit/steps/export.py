"""`toolkit export` — one xlsx of everything produced so far.

Reads the deliverables under outputs/ and writes outputs/export.xlsx with three tabs:
- Clips: one row per clip (id, interview, session, start/end, label, per-topic-set tags,
  locations, regions);
- Interviews: one row per narrator (sessions, summary, per-topic-set tags, locations);
- Categories: the vocabularies (each topic set's names, the location labels) as reference columns.

How locations appear is `export.locations` in config.yaml — see LOCATION_MODES.

Incremental: a column appears only if its step has run; missing steps are announced, not fatal.
Overwrites the file each run (idempotent). No live Google Sheets — this produces a plain xlsx
you can open in Excel or upload to Google Sheets.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..core.config import load_root_config, load_step_config
from ..core.ids import narrator_key
from ..errors import ToolkitError
from ..project import Project
from .topics.taxonomy import load_topic_set

# How location tags land in the spreadsheet. The tagger records countries and regions separately,
# and `toolkit locations map` expands each region into its countries; these modes pick which of
# those views the deliverable shows.
LOCATION_MODES = {
    "countries": "only the countries tagged directly",
    "countries_and_regions": "those countries, plus a separate Regions column",
    "countries_incl_regions": "one column: direct countries plus regions mapped down to countries",
}
DEFAULT_LOCATION_MODE = "countries_and_regions"
DEFAULT_FILENAME = "export.xlsx"


def _location_mode(project: Project, override: str | None = None) -> str:
    mode = override or (load_root_config(project).get("export") or {}).get(
        "locations", DEFAULT_LOCATION_MODE)
    if mode not in LOCATION_MODES:
        raise ToolkitError(
            f"export.locations must be one of {', '.join(LOCATION_MODES)} (got {mode!r}).")
    return mode


def _read(path):
    return pd.read_parquet(path) if path.exists() else None


def _direct_only(long_df: pd.DataFrame) -> pd.DataFrame:
    """Rows whose label was tagged directly (or is a configured place-tag), dropping labels that
    exist only because a region was expanded into them. Same provenance rule the rollup uses."""
    mask = long_df["via"].str.split("|").map(
        lambda v: "direct" in v or "place" in v).astype(bool)
    return long_df[mask]


def _joined(df: pd.DataFrame, key: str, value: str) -> dict[str, str]:
    """{key: "a, b, c"} from a long table, labels sorted for a stable deliverable."""
    return {k: ", ".join(sorted(g[value])) for k, g in df.groupby(key)}


def _topic_sets(project: Project) -> list[str]:
    topics = (load_root_config(project).get("topics") or {})
    return list((topics.get("sets") or {}).keys())


def _clip_topic_tags(project: Project, set_name: str) -> dict[str, str] | None:
    """clip_id -> comma-joined names of that set's assigned (score==max) topics."""
    long = _read(project.outputs_dir / "topics" / f"{set_name}_clip_topics_long.parquet")
    if long is None:
        return None
    top = long["score"].max() if len(long) else 0
    assigned = long[long["score"] == top]
    return {cid: ", ".join(sorted(g["topic_name"])) for cid, g in assigned.groupby("clip_id")}


def build_clips_sheet(project: Project, sets: list[str],
                      location_mode: str = DEFAULT_LOCATION_MODE) -> tuple[pd.DataFrame, list[str]]:
    labels = _read(project.outputs_dir / "labels" / "labels.parquet")
    clips = labels if labels is not None else _read(project.outputs_dir / "clips" / "clips.parquet")
    if clips is None:
        raise ToolkitError("No clips yet — run `toolkit clip` first (export needs at least clips).")

    session_regex = load_step_config(project, "import")["session_regex"]
    df = pd.DataFrame({
        "Clip Id": clips["clip_id"],
        "Interview": clips["interview_id"].map(lambda i: narrator_key(i, session_regex)),
        "Session": clips["interview_id"],
        "Start": clips["start_ts"],
        "End": clips["end_ts"],
    })
    included = ["clips"]
    if labels is not None:
        from ..core import overrides as overrides_mod

        edits, complaints = overrides_mod.overlay(project, clips)
        for reason in complaints:
            print(f"⚠ {reason}")
        df["Label"] = [edits.get(cid, lab)
                       for cid, lab in zip(clips["clip_id"], clips["label"])]
        if edits:
            print(f"{len(edits)} label(s) are edited by hand (label_overrides.csv) — the sheet "
                  f"shows your versions.")
        included.append("labels")

    for set_name in sets:
        tags = _clip_topic_tags(project, set_name)
        if tags is not None:
            df[f"Topics: {set_name}"] = df["Clip Id"].map(tags).fillna("")
            included.append(f"topics:{set_name}")

    countries = _read(project.outputs_dir / "locations" / "clip_countries.parquet")
    if countries is not None:
        if location_mode == "countries_incl_regions":
            cmap = dict(zip(countries["clip_id"],
                            countries["countries_final"].str.replace("|", ", ")))
        else:                       # direct tags only — regions are not folded in
            long = _read(project.outputs_dir / "locations" / "clip_countries_long.parquet")
            cmap = _joined(_direct_only(long), "clip_id", "country") if long is not None else {}
        df["Locations"] = df["Clip Id"].map(cmap).fillna("")
        if location_mode == "countries_and_regions":
            rmap = dict(zip(countries["clip_id"], countries["regions"].str.replace("|", ", ")))
            df["Regions"] = df["Clip Id"].map(rmap).fillna("")
        included.append("locations")
    return df, included


def build_interviews_sheet(project: Project, sets: list[str],
                           location_mode: str = DEFAULT_LOCATION_MODE) -> pd.DataFrame | None:
    session_regex = load_step_config(project, "import")["session_regex"]
    frames: dict[str, dict] = {}

    def row(key: str) -> dict:
        return frames.setdefault(key, {"Interview": key})

    summaries = _read(project.outputs_dir / "summaries" / "summaries.parquet")
    if summaries is not None:
        # A transcript that was never SYNC'd can only be summarized, so its row is a summary and
        # nothing else. The column appears only where there is one, and it is what tells a reader
        # that the empty tags are a fact about the transcript rather than a gap in the work.
        some_unsynced = "synced" in summaries.columns and not summaries["synced"].all()
        for r in summaries.itertuples():
            rr = row(r.interview_key)
            rr["Sessions"] = str(r.session_ids).replace("|", ", ")
            rr["Summary"] = r.summary
            if some_unsynced:
                rr["Transcript"] = "SYNC'd" if getattr(r, "synced", True) else "not SYNC'd"

    for set_name in sets:
        wide = _read(project.outputs_dir / "topics" / f"{set_name}_interview_topics_wide.parquet")
        if wide is not None:
            for r in wide.itertuples():
                row(r.interview_key)[f"Topics: {set_name}"] = str(r.topics).replace("|", ", ")

    loc = _read(project.outputs_dir / "locations" / "interview_locations_wide.parquet")
    if loc is not None:
        direct = None
        if location_mode != "countries_incl_regions":
            long = _read(project.outputs_dir / "locations" / "interview_locations_long.parquet")
            direct = _joined(_direct_only(long), "interview_key", "label") if long is not None else {}
        for r in loc.itertuples():
            rr = row(r.interview_key)
            rr["Locations"] = (str(r.labels).replace("|", ", ") if direct is None
                               else direct.get(r.interview_key, ""))
            if location_mode == "countries_and_regions":
                rr["Regions"] = str(r.regions).replace("|", ", ")

    if not frames:
        return None

    # When each narrator's transcripts were imported — one timestamp per session, aligned with
    # the Sessions cell. A corrected transcript re-imported later carries a newer stamp, which
    # is how a reader checks a sheet against the current text ("exported before that → redo").
    stamps = _imported_stamps(project, session_regex)
    for key, rr in frames.items():
        if key in stamps:
            rr["Imported"] = stamps[key]

    # a Session column derived from clips if summaries didn't populate one
    df = pd.DataFrame(list(frames.values()))
    return df.sort_values("Interview").reset_index(drop=True)


def _imported_stamps(project: Project, session_regex: str) -> dict[str, str]:
    """narrator -> "2026-08-06 14:32" (comma-joined per session, in session id order)."""
    from ..core import manifest as manifest_mod

    stamps: dict[str, dict[str, str]] = {}
    for pile in (manifest_mod.SYNCED, manifest_mod.UNSYNCED):
        for iid, iso in manifest_mod.imported_at(project, pile).items():
            stamps.setdefault(narrator_key(iid, session_regex), {})[iid] = iso
    return {key: ", ".join(manifest_mod.local_stamp(by_id[i]) for i in sorted(by_id))
            for key, by_id in stamps.items()}


def build_categories_sheet(project: Project, sets: list[str],
                           location_mode: str = DEFAULT_LOCATION_MODE) -> pd.DataFrame:
    cfg = load_root_config(project)
    columns: dict[str, list[str]] = {}
    for set_name in sets:
        try:
            ts = load_topic_set(project, load_step_config(project, "topics"), set_name)
            columns[f"Topics: {set_name}"] = [t["name"] for t in ts.topics]
        except ToolkitError:
            continue
    # The reference lists must match what the other tabs actually contain, or they invite
    # filtering by a value that appears nowhere.
    if location_mode == "countries_and_regions":
        regions_file = load_step_config(project, "locations").get("regions_file")
        if regions_file:
            path = project.root / regions_file
            if path.exists():
                import yaml
                regions = yaml.safe_load(path.read_text()) or []
                columns["Regions"] = list(regions)
    countries = _read(project.outputs_dir / "locations" / "clip_countries_long.parquet")
    if countries is not None and len(countries):
        shown = countries if location_mode == "countries_incl_regions" else _direct_only(countries)
        if len(shown):
            columns["Locations"] = sorted(shown["country"].unique())
    if not columns:
        return pd.DataFrame()
    width = max(len(v) for v in columns.values())
    return pd.DataFrame({k: v + [""] * (width - len(v)) for k, v in columns.items()})


def _write_export_manifest(project: Project, clips_df: pd.DataFrame) -> None:
    """What this export said each label was — the reference the next export diffs the sheet
    against. A cell that differs from it was edited by a person; a cell that matches was not,
    even if the pipeline has re-labeled since."""
    import json

    if "Label" not in clips_df.columns:
        return
    project.export_manifest_path.parent.mkdir(parents=True, exist_ok=True)
    project.export_manifest_path.write_text(json.dumps({
        "schema": 1,
        "labels": dict(zip(clips_df["Clip Id"].astype(str), clips_df["Label"].astype(str))),
    }, indent=2) + "\n")


def _harvest_sheet_edits(project: Project, out_path: Path, clips_tab: str) -> int:
    """Labels edited in the exported sheet, upserted into label_overrides.csv.

    Only a cell that differs from what the LAST export wrote counts as an edit; without that
    reference, a regenerated label would be indistinguishable from a hand-fix. Editing a cell
    back to the model's own words removes the override again."""
    import json

    if not out_path.exists() or not project.export_manifest_path.exists():
        return 0
    last = (json.loads(project.export_manifest_path.read_text()).get("labels")) or {}
    labels = _read(project.outputs_dir / "labels" / "labels.parquet")
    if not last or labels is None:
        return 0

    from openpyxl import load_workbook
    try:
        wb = load_workbook(out_path, read_only=True, data_only=True)
    except Exception as e:
        raise ToolkitError(
            f"Could not read the existing {out_path.name}, so labels edited in the sheet could "
            f"not be kept. Close it if it is open in Excel, or delete it if there is nothing "
            f"in it to keep. ({e})") from e
    if clips_tab not in wb.sheetnames:
        wb.close()
        return 0

    from ..core import overrides as overrides_mod

    rows = wb[clips_tab].iter_rows(values_only=True)
    header = list(next(rows, ()) or ())
    if "Clip Id" not in header or "Label" not in header:
        wb.close()
        return 0
    id_col, label_col = header.index("Clip Id"), header.index("Label")
    model_label = dict(zip(labels["clip_id"].astype(str), labels["label"].astype(str)))
    kept = 0
    for row in rows:
        cid = None if row[id_col] is None else str(row[id_col])
        if not cid or cid not in model_label:
            continue                    # a clip this pipeline no longer knows — nothing to pin
        sheet_label = "" if row[label_col] is None else str(row[label_col])
        exported = last.get(cid)
        if exported is None or sheet_label == exported or not sheet_label.strip():
            continue                    # untouched (or emptied, which is not a label)
        if sheet_label == model_label[cid]:
            if overrides_mod.remove(project, cid):      # edited back to the model's own words
                kept += 1
            continue
        overrides_mod.upsert(project, cid, sheet_label, labels, replaces=model_label[cid])
        kept += 1
    wb.close()
    return kept


def _write_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title)
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append(["" if pd.isna(v) else v for v in r.tolist()])
    for i, col in enumerate(df.columns, start=1):
        longest = max([len(str(col))] + [len(str(v)) for v in df[col].tolist()[:200]], default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 60)


def run_export(project: Project, out: str | None = None, locations: str | None = None) -> None:
    cfg = load_step_config(project, "export")
    sets = _topic_sets(project)
    mode = _location_mode(project, locations)
    tabs = cfg.get("tabs") or {}
    out_path = (project.root / out) if out else (project.outputs_dir /
                                                 cfg.get("filename", DEFAULT_FILENAME))

    # Before the old sheet is overwritten: labels edited in it become overrides, so a re-export
    # never silently undoes a curator's hand-fix.
    kept = _harvest_sheet_edits(project, out_path, tabs.get("clips", "Clips"))

    clips_df, included = build_clips_sheet(project, sets, mode)
    interviews_df = build_interviews_sheet(project, sets, mode)
    categories_df = build_categories_sheet(project, sets, mode)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, tabs.get("clips", "Clips"), clips_df)
    if interviews_df is not None:
        _write_sheet(wb, tabs.get("interviews", "Interviews"), interviews_df)
    if not categories_df.empty:
        _write_sheet(wb, tabs.get("categories", "Categories"), categories_df)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    _write_export_manifest(project, clips_df)

    print(f"Wrote {out_path}")
    if kept:
        print(f"  Kept {kept} label(s) you edited in the sheet -> "
              f"{project.label_overrides_path.name}")
    print(f"  Clips tab: {len(clips_df)} clips, columns include: {', '.join(included)}")
    if interviews_df is not None:
        print(f"  Interviews tab: {len(interviews_df)} narrators")
    if "locations" in included:
        print(f"  Locations: {mode} — {LOCATION_MODES[mode]}")
    all_steps = {"clips", "labels", "locations"} | {f"topics:{s}" for s in sets}
    missing = sorted(all_steps - set(included))
    if missing:
        print(f"  Not yet included (step not run): {', '.join(missing)}")
