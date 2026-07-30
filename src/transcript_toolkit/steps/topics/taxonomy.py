"""Topic sets loaded from user spreadsheets (config: topics.sets.<set>.file).

The research repo kept each taxonomy as hand-edited markdown plus a config-listed id/name
table; the product replaces both with ONE spreadsheet per set — columns `name` (required),
`description` (required), `id` (optional; else slugified from the name). The loader
deterministically generates everything the tagger needs: the taxonomy text fed to the model,
the ordered [{id, name}] list, and (via `build_legend`) the topic-id legend.

BYTE-STABILITY WARNING: the generated taxonomy text and legend feed cache keys and demo
fingerprints. Any cosmetic change to the generated format invalidates users' caches and
recorded demos.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from ...core import thresholds
from ...errors import ToolkitError
from ...project import Project

ID_RE = re.compile(r"^[a-z0-9_]+$")

# What a topic list gets when it is first used: the toolkit's own default rollup rule, written
# out so the setting is visible in config.yaml rather than implied by its absence.
DEFAULT_ROLLUP = thresholds.DEFAULT.as_config()
ROLLUP_LINE = "{ " + ", ".join(f"{k}: {v}" for k, v in DEFAULT_ROLLUP.items()) + " }"


@dataclass(frozen=True)
class TopicSet:
    name: str                # the set's config name, e.g. "main" — prefixes outputs and state
    ids: list[str]
    topics: list[dict]       # ordered [{id, name}], spreadsheet row order
    taxonomy_text: str       # deterministic markdown fed to the model (feeds cache keys)
    source: Path
    prompt: str | None       # optional per-set prompt file (config sets.<set>.prompt);
                             # None = the step-wide advanced `prompt` is used
    overrides: dict          # settings this set runs with instead of the step's (SET_OVERRIDES)
    rollup: thresholds.Rollup  # how this list's clip tags become interview tags


# A topic list is its own piece of work: a fine-grained list may want a stronger model or more
# thinking than a coarse one, and there is no reason for one to dictate the other. A set that
# overrides nothing runs with the step's own settings.
SET_OVERRIDES = ("model", "reasoning")


TOPIC_SUFFIXES = (".csv", ".xlsx")
EXAMPLE_STEM = "example_topics"          # the shipped template; not a usable set on its own


def unusable_set_name(name: str) -> str | None:
    """Why this cannot be a set name, or None when it can.

    The set name becomes a key under `topics.sets`, and settings address that key by a dotted
    path (`topics.sets.collection.rollup`). A dot in the name would split there and write the
    rollup rule into a level nothing reads — so a list called `themes.v2.csv` is refused by name
    rather than saved into the wrong place.
    """
    return "it has a dot in it" if "." in name else None


def discover_topic_files(project: Project) -> dict[str, Path]:
    """{set name: spreadsheet} for every topic list sitting in topics/, the set name being the
    filename without its extension. The shipped example template is excluded — it is meant to be
    filled in and renamed, not tagged against."""
    found: dict[str, Path] = {}
    if not project.topics_dir.is_dir():
        return found
    for path in sorted(project.topics_dir.iterdir()):
        if (path.suffix.lower() in TOPIC_SUFFIXES and path.stem != EXAMPLE_STEM
                and not unusable_set_name(path.stem)):
            found.setdefault(path.stem, path)     # .csv wins over .xlsx for the same stem
    return found


def unusable_topic_files(project: Project) -> dict[str, str]:
    """{filename: why it is not offered as a set}. Skipping one silently would leave somebody
    looking at a list in topics/ that nothing anywhere admits to seeing."""
    if not project.topics_dir.is_dir():
        return {}
    return {path.name: why
            for path in sorted(project.topics_dir.iterdir())
            if path.suffix.lower() in TOPIC_SUFFIXES and path.stem != EXAMPLE_STEM
            and (why := unusable_set_name(path.stem))}


def available_sets(project: Project, cfg: dict) -> list[str]:
    """Every set a user could name: configured ones plus spreadsheets waiting in topics/."""
    configured = cfg.get("sets") or {}
    names = set(configured) if isinstance(configured, dict) else set()
    return sorted(names | set(discover_topic_files(project)))


def _skipped_note(project: Project) -> str:
    skipped = unusable_topic_files(project)
    if not skipped:
        return ""
    listed = "; ".join(f"{name} ({why})" for name, why in skipped.items())
    return (f"\nNot usable as a topic list: {listed}. Rename the file — the name without the "
            f"extension becomes the set name.")


def _no_set_error(project: Project, cfg: dict, given: str | None) -> ToolkitError:
    names = available_sets(project, cfg)
    lead = (f"Unknown topic set {given!r}. There is no topics.sets.{given} in config.yaml and no "
            f"topics/{given}.csv or topics/{given}.xlsx." if given
            else "No topic set given. Use --set <name>.")
    skipped = _skipped_note(project)
    if names:
        return ToolkitError(f"{lead}\nAvailable: {', '.join(names)}{skipped}")
    example = project.topics_dir / f"{EXAMPLE_STEM}.csv"
    hint = (f"\nStart from {example.name}: fill in your topics, then rename it to the set name you "
            f"want (e.g. topics/collection.csv)." if example.exists() else "")
    return ToolkitError(
        f"{lead}\nNo topic lists found. Put one in {project.topics_dir}{'/'} as a .csv or .xlsx — "
        f"the filename becomes the set name.{skipped}{hint}\n"
        f"Then run: toolkit topics tag --set <name> --demo")


def resolve_set(project: Project, cfg: dict, set_name: str | None) -> tuple[str, dict]:
    """Resolve `--set NAME` to (name, its config entry).

    A set may simply be a spreadsheet dropped into topics/ — no config editing needed. The first
    time such a set is used it is written into config.yaml (so its rollup and any per-set prompt
    are visible and editable afterwards). There is deliberately no default set: tagging the wrong
    taxonomy is expensive, so the set is always named explicitly."""
    sets = cfg.get("sets") or {}
    if not isinstance(sets, dict):
        raise ToolkitError("config.yaml topics.sets must be a mapping of set name -> settings.")
    if not set_name:
        raise _no_set_error(project, cfg, None)
    why = unusable_set_name(set_name)
    if why:
        raise ToolkitError(
            f"{set_name!r} cannot be a topic set name: {why}. The name is a settings key, and a "
            f"dot in it would put this list's settings somewhere nothing reads them. Rename the "
            f"spreadsheet in topics/ (and its entry in config.yaml, if it has one).")

    discovered = discover_topic_files(project)
    if set_name in sets:
        entry = sets[set_name]
        if not isinstance(entry, dict):
            raise ToolkitError(f"config.yaml topics.sets.{set_name} must be a mapping")
        if entry.get("file") or set_name not in discovered:
            return set_name, entry
        # An entry with settings but no spreadsheet named. That is what config.yaml looks like
        # after somebody sets a list's rollup rule before ever tagging with it — the app offers
        # that from the start — and refusing here would break a list that is sitting in topics/
        # in plain sight. The filename is the set name, which is the whole convention.
        return set_name, {**entry,
                          "file": discovered[set_name].relative_to(project.root).as_posix()}

    if set_name not in discovered:
        raise _no_set_error(project, cfg, set_name)

    file_rel = discovered[set_name].relative_to(project.root).as_posix()
    entry = {"file": file_rel, "rollup": dict(DEFAULT_ROLLUP)}
    if register_topic_set(project, set_name, file_rel):
        print(f"Registered topic set '{set_name}' in config.yaml (file: {file_rel}, "
              f"rollup: {thresholds.DEFAULT.describe()}). Run "
              f"`toolkit topics thresholds --set {set_name}` to see what the alternatives "
              f"would tag, and edit the rollup there to change it.")
    else:
        print(f"NOTE: could not add '{set_name}' to config.yaml automatically — using "
              f"file: {file_rel} with the default rollup for this run.\n"
              f"      To make it permanent, add under `topics:` -> `sets:`:\n"
              f"        {set_name}:\n          file: {file_rel}\n"
              f"          rollup: {ROLLUP_LINE}")
    return set_name, entry


def register_topic_set(project: Project, name: str, file_rel: str) -> bool:
    """Append topics.sets.<name> to config.yaml as TEXT, so the file's comments (which are the
    user-facing documentation of every setting) survive — a yaml load/dump round-trip would strip
    them all. Returns False, changing nothing, if the file does not have the shape we expect or
    the edit would alter anything other than adding this one set."""
    import yaml

    path = project.config_path
    text = path.read_text()
    try:
        before = yaml.safe_load(text) or {}
    except yaml.YAMLError:
        return False
    if not isinstance(before, dict):
        return False
    topics_before = before.get("topics") or {}
    sets_before = topics_before.get("sets") or {} if isinstance(topics_before, dict) else None
    if sets_before is None or not isinstance(sets_before, dict) or name in sets_before:
        return False

    block = [f"    {name}:", f"      file: {file_rel}", f"      rollup: {ROLLUP_LINE}"]
    lines = text.splitlines()

    def top_level_end(start: int) -> int:
        """Index just past the block owned by the top-level key opened at `start`."""
        for i in range(start + 1, len(lines)):
            if lines[i].strip() and not lines[i][0].isspace():
                return i
        return len(lines)

    topics_at = next((i for i, ln in enumerate(lines) if ln.rstrip() == "topics:"), None)
    if topics_at is None:                                   # no topics block at all: append one
        new_lines = lines + ["", "topics:", "  sets:", *block]
    else:
        end = top_level_end(topics_at)
        sets_at = next((i for i in range(topics_at + 1, end)
                        if lines[i].rstrip() in ("  sets:", "  sets: {}")), None)
        if sets_at is None:                                 # topics block without a sets: key
            new_lines = lines[:end] + ["  sets:", *block] + lines[end:]
        else:
            new_lines = lines[:sets_at + 1] + block + lines[sets_at + 1:]
            if lines[sets_at].rstrip() == "  sets: {}":
                new_lines[sets_at] = "  sets:"

    new_text = "\n".join(new_lines) + ("\n" if text.endswith("\n") else "")
    try:                                                    # the edit must do exactly one thing
        after = yaml.safe_load(new_text) or {}
    except yaml.YAMLError:
        return False
    expected = {**before, "topics": {**topics_before,
                                     "sets": {**sets_before,
                                              name: {"file": file_rel,
                                                     "rollup": dict(DEFAULT_ROLLUP)}}}}
    if after != expected:
        return False
    path.write_text(new_text)
    return True


def slug(name: str) -> str:
    """Topic id from a display name: lowercase, runs of non-alphanumerics -> _."""
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def read_table(path: Path) -> list[list[str]]:
    """Raw rows (header first) from a .csv or .xlsx — every cell stringified, deterministic."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [["" if c is None else str(c) for c in row] for row in csv.reader(f)]
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return [["" if c is None else str(c) for c in row]
                    for row in wb.worksheets[0].iter_rows(values_only=True)]
        finally:
            wb.close()
    raise ToolkitError(f"Topic set file must be .csv or .xlsx, got: {path.name}")


def load_topic_set(project: Project, cfg: dict, set_name: str | None = None) -> TopicSet:
    """Load one topic set's spreadsheet into a TopicSet. `cfg` is the merged topics step
    config (load_step_config(project, "topics")); the `sets` dict arrives nested in it."""
    name, entry = resolve_set(project, cfg, set_name)
    file_rel = entry.get("file")
    if not file_rel:
        raise ToolkitError(f"config.yaml topics.sets.{name} has no `file` "
                           f"(path to the topic spreadsheet, relative to the workspace).")
    path = project.root / file_rel
    if not path.exists():
        raise ToolkitError(f"Topic set file not found: {path}")

    rows, blocks = read_topic_rows(read_table(path), path.name)
    # TopicSet carries id and name only: the descriptions are already in taxonomy_text, which
    # is what feeds the model and the cache key.
    topics = [{"id": r["id"], "name": r["name"]} for r in rows]
    return TopicSet(name=name, ids=[t["id"] for t in topics], topics=topics,
                    taxonomy_text="\n\n".join(blocks), source=path,
                    prompt=entry.get("prompt"),
                    overrides={k: entry[k] for k in SET_OVERRIDES if entry.get(k) is not None},
                    rollup=thresholds.parse(entry.get("rollup"),
                                            f"config.yaml topics.sets.{name}.rollup"))


def read_topic_rows(raw: list[list[str]], label: str) -> tuple[list[dict], list[str]]:
    """Validate a topic table and return ([{id, name, description}], taxonomy blocks).

    Split out of `load_topic_set` so the app can check a list someone is typing against exactly
    the rules the run will apply, and complain in exactly the same words — a topic list that the
    editor accepted and the step then rejected would be the worst of both.
    """
    if not raw:
        raise ToolkitError(f"{label} is empty")
    header = [h.strip().lower() for h in raw[0]]
    for col in ("name", "description"):
        if col not in header:
            raise ToolkitError(f"{label} needs a {col!r} column "
                               f"(found: {', '.join(h for h in header if h) or '(none)'})")

    topics: list[dict] = []
    blocks: list[str] = []
    seen: dict[str, int] = {}          # id -> row number, for duplicate reporting
    for rownum, cells in enumerate(raw[1:], start=2):
        row = dict(zip(header, cells))
        if not any(v.strip() for v in row.values()):
            continue                   # blank row (common in xlsx exports)
        topic_name = (row.get("name") or "").strip()
        description = (row.get("description") or "").strip()
        if not topic_name:
            raise ToolkitError(f"{label} row {rownum}: empty topic name")
        if not description:
            raise ToolkitError(f"{label} row {rownum}: empty description for topic {topic_name!r}")
        tid = (row.get("id") or "").strip() or slug(topic_name)
        if not ID_RE.match(tid):
            raise ToolkitError(f"{label} row {rownum}: invalid topic id {tid!r} "
                               f"(must match ^[a-z0-9_]+$; give an explicit `id` column value)")
        if tid in seen:
            raise ToolkitError(f"{label} row {rownum}: duplicate topic id {tid!r} "
                               f"(also produced by row {seen[tid]})")
        seen[tid] = rownum
        topics.append({"id": tid, "name": topic_name, "description": description})
        blocks.append(f"## {topic_name}\n\n{description}")
    if not topics:
        raise ToolkitError(f"{label} has no topic rows")
    return topics, blocks


def build_legend(topics: list[dict]) -> str:
    """The id<->name legend prepended to the taxonomy so the model knows which id to use.
    Ported byte-identical from the working repo's tag-topics utils."""
    lines = ["## Topics", "",
             "Score the clip against each of these topics, using exactly these ids in your output. "
             "Definitions follow below.", ""]
    lines += [f"- `{t['id']}` — {t['name']}" for t in topics]
    return "\n".join(lines)
