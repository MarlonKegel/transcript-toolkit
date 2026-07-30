"""Reading and writing topic lists from the app.

A topic list is a spreadsheet in `topics/` — that is the whole format, and the app must not
invent a second one. So the editor is a view onto that file: it reads the same table the run
reads, checks it with the run's own rules (`read_topic_rows`), and writes plain CSV back.

Until a list has been named it is written to the shipped example file. That file is deliberately
excluded from set discovery, so half-finished editing can never be tagged against by accident.
"""
from __future__ import annotations

import csv
from pathlib import Path

from ..errors import ToolkitError
from ..project import Project
from ..steps.topics.taxonomy import (EXAMPLE_STEM, read_table, read_topic_rows,
                                     register_topic_set, slug)

COLUMNS = ("id", "name", "description")     # fixed: the run reads these names
HOW_TO_ROW = "DELETE THIS ROW"              # the example file's own instructions, as a row


def draft_path(project: Project) -> Path:
    """Where an unnamed list is kept: the shipped example, which is never a usable set."""
    return project.topics_dir / f"{EXAMPLE_STEM}.csv"


def set_path(project: Project, name: str) -> Path:
    return project.topics_dir / f"{name}.csv"


def discovered_path(project: Project, name: str) -> Path | None:
    """Where a list that was simply dropped into topics/ actually is — it may be an .xlsx, and
    it is a set whether or not config.yaml has caught up with it yet."""
    from ..steps.topics.taxonomy import discover_topic_files
    return discover_topic_files(project).get(name)


def load_rows(path: Path) -> tuple[list[dict], str]:
    """The list as rows, plus the guidance the example file carries in its first row.

    Returns rows even when the list would not pass validation — this is an editor, and someone
    with a half-finished list has to be able to see it and fix it.
    """
    if not path.exists():
        return [], ""
    raw = read_table(path)
    if not raw:
        return [], ""
    header = [h.strip().lower() for h in raw[0]]
    rows, guidance = [], ""
    for cells in raw[1:]:
        row = {k: (v or "").strip() for k, v in zip(header, cells)}
        if not any(row.values()):
            continue
        if row.get("id") == HOW_TO_ROW:
            guidance = row.get("description", "")
            continue
        rows.append({c: row.get(c, "") for c in COLUMNS})
    return rows, guidance


def check(rows: list[dict], label: str) -> None:
    """Refuse a list the run would refuse, in the run's own words."""
    read_topic_rows(as_table(rows), label)


def as_table(rows: list[dict]) -> list[list[str]]:
    return [list(COLUMNS)] + [[row.get(c, "") or "" for c in COLUMNS] for row in rows]


def write_rows(path: Path, rows: list[dict]) -> None:
    """Write the list back to the file it came from. Blank rows are dropped — an editor
    collects them.

    An .xlsx keeps being an .xlsx: somebody who brought a spreadsheet gets to keep the format
    they brought, and any other sheet in the workbook is left where it is. Only the sheet the
    toolkit reads is rewritten, so formatting on that one sheet does not survive.
    """
    kept = [r for r in rows if any((r.get(c) or "").strip() for c in COLUMNS)]
    table = [list(COLUMNS)] + [[(row.get(c) or "").strip() for c in COLUMNS] for row in kept]
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".xlsx":
        _write_xlsx(path, table)
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for line in table:
            writer.writerow(line)


def _write_xlsx(path: Path, table: list[list[str]]) -> None:
    from openpyxl import Workbook, load_workbook

    if path.exists():
        book = load_workbook(path)
        sheet = book.active
        book.remove(sheet)
        sheet = book.create_sheet(sheet.title, 0)
    else:
        book = Workbook()
        sheet = book.active
    for line in table:
        sheet.append(line)
    book.save(path)


def save_draft(project: Project, rows: list[dict]) -> Path:
    """Keep what has been typed so far, without making it a set anybody can run."""
    path = draft_path(project)
    write_rows(path, rows)
    return path


def valid_set_name(name: str) -> str:
    """A set name is a filename, so it has to survive being one."""
    cleaned = slug(name.strip())
    if not cleaned:
        raise ToolkitError(f"{name!r} cannot be a topic list name. Use letters and numbers — "
                           f"for example: collection, filter, themes.")
    return cleaned


def save_as(project: Project, name: str, rows: list[dict]) -> tuple[str, Path]:
    """Name the list, write it, and make it a set the steps can use.

    Registering it in config.yaml is what turns a spreadsheet into something `--set` accepts,
    and it is the same edit `toolkit topics tag` would offer to make.
    """
    set_name = valid_set_name(name)
    path = set_path(project, set_name)
    if path.exists():
        raise ToolkitError(f"There is already a topic list called {set_name!r} "
                           f"({path.name}). Pick a different name.")
    check(rows, path.name)
    write_rows(path, rows)
    register_topic_set(project, set_name, f"topics/{path.name}")
    return set_name, path


def save_existing(project: Project, set_name: str, path: Path, rows: list[dict]) -> None:
    """Overwrite a named list that is already in use, in whatever format it is kept in."""
    check(rows, path.name)
    write_rows(path, rows)
